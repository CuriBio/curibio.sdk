# -*- coding: utf-8 -*-
"""Classes and functions for finding and managing excel files."""
import datetime
from typing import Any
from typing import Optional
from uuid import UUID

from mantarray_file_manager import CURI_BIO_ACCOUNT_UUID
from mantarray_file_manager import CURI_BIO_USER_ACCOUNT_ID
from mantarray_file_manager import MANTARRAY_SERIAL_NUMBER_UUID
from mantarray_file_manager import PLATE_BARCODE_UUID
from mantarray_file_manager import TISSUE_SAMPLING_PERIOD_UUID
from mantarray_file_manager import UTC_BEGINNING_RECORDING_UUID
from mantarray_file_manager import WELL_NAME_UUID
from mantarray_file_manager import WellFile
from nptyping import NDArray
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from stdlib_utils import get_current_file_abs_directory
from xlsxwriter.utility import xl_cell_to_rowcol

from .constants import EXCEL_OPTICAL_METADATA_CELLS
from .constants import TWITCHES_POINT_UP_UUID

PATH_OF_CURRENT_FILE = get_current_file_abs_directory()


def _get_col_as_array(
    sheet: Worksheet,
    zero_based_row: int,
    zero_based_col: int,
) -> NDArray[(2, Any), float]:
    result = _get_cell_value(sheet, zero_based_row, zero_based_col)
    col_array = [result]
    zero_based_row += 1
    while result:
        result = _get_cell_value(sheet, zero_based_row, zero_based_col)
        col_array.append(result)
        zero_based_row += 1
    return np.array(col_array)


def _get_well_index_from_well_name(well_name: str) -> int:
    return (ord(well_name[0]) - ord("A")) * 4 + int(well_name[1]) - 1


def _get_single_sheet(file_name: str) -> Any:
    wb = load_workbook(file_name)
    return wb[wb.sheetnames[0]]


def _get_cell_value(
    sheet: Worksheet,
    zero_based_row: int,
    zero_based_col: int,
) -> Optional[str]:
    result = sheet.cell(row=zero_based_row + 1, column=zero_based_col + 1).value
    if result is None:
        return result
    return str(result)


def _get_excel_metadata_value(sheet: Worksheet, metadata_uuid: UUID) -> str:
    cell_name = EXCEL_OPTICAL_METADATA_CELLS[metadata_uuid]
    row, col = xl_cell_to_rowcol(cell_name)
    result = _get_cell_value(sheet, row, col)
    if result is None:
        raise Exception()  # raise some custom error here: MetadataNotFoundError
    return result


class ExcelWellFile(WellFile):
    """Wrapper around an Excel file for a single well of optical data.

    Args:
        file_name: The path of the excel file to open.

    Attributes:
        _excel_sheet: The opened excel sheet.
    """

    def __init__(self, file_name: str) -> None:
        self._excel_sheet = _get_single_sheet(file_name)
        self._file_name = file_name
        self._file_version = "0.1.0"
        self._raw_tissue_reading: Optional[NDArray[(2, Any), int]] = None
        self._raw_ref_reading: Optional[NDArray[(2, Any), int]] = None

    def get_h5_file(self) -> None:
        raise NotImplementedError("ExcelWellFiles do not store an H5 file")

    def get_h5_attribute(self, attr_name: str) -> Any:
        raise NotImplementedError(
            "ExcelWellFiles do not store an H5 file and therefore cannot get H5 attributes"
        )

    def get_well_name(self) -> str:
        return _get_excel_metadata_value(self._excel_sheet, WELL_NAME_UUID)

    def get_well_index(self) -> int:
        return _get_well_index_from_well_name(self.get_well_name())

    def get_plate_barcode(self) -> str:
        return _get_excel_metadata_value(self._excel_sheet, PLATE_BARCODE_UUID)

    def get_user_account(self) -> UUID:
        if not isinstance(CURI_BIO_USER_ACCOUNT_ID, UUID):
            # Tanner (10/13/20): Making mypy happy
            raise NotImplementedError(
                "CURI_BIO_USER_ACCOUNT_ID should always be a UUID"
            )
        return CURI_BIO_USER_ACCOUNT_ID

    def get_timestamp_of_beginning_of_data_acquisition(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_customer_account(self) -> UUID:
        if not isinstance(CURI_BIO_ACCOUNT_UUID, UUID):
            # Tanner (10/13/20): Making mypy happy
            raise NotImplementedError("CURI_BIO_ACCOUNT_UUID should always be a UUID")
        return CURI_BIO_ACCOUNT_UUID

    def get_mantarray_serial_number(self) -> str:
        return _get_excel_metadata_value(
            self._excel_sheet, MANTARRAY_SERIAL_NUMBER_UUID
        )

    def get_begin_recording(self) -> datetime.datetime:
        timestamp_str = _get_excel_metadata_value(
            self._excel_sheet, UTC_BEGINNING_RECORDING_UUID
        )
        timestamp = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        return timestamp

    def get_timestamp_of_first_tissue_data_point(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_timestamp_of_first_ref_data_point(self) -> datetime.datetime:
        return self.get_begin_recording()

    def get_tissue_sampling_period_microseconds(self) -> int:
        sampling_period_seconds = float(
            _get_excel_metadata_value(self._excel_sheet, TISSUE_SAMPLING_PERIOD_UUID)
        )
        return int(round(sampling_period_seconds, 6) * 1e6)

    def get_reference_sampling_period_microseconds(self) -> int:
        return 0

    def get_recording_start_index(self) -> int:
        return 0

    def get_twitches_point_up(self) -> bool:
        return "y" in _get_excel_metadata_value(
            self._excel_sheet, TWITCHES_POINT_UP_UUID
        )

    def get_raw_tissue_reading(self) -> NDArray[(2, Any), float]:
        if self._raw_tissue_reading is None:
            self._raw_tissue_reading = np.array(
                (
                    _get_col_as_array(self._excel_sheet, 1, 0),
                    _get_col_as_array(self._excel_sheet, 1, 1),
                )
            )
        return self._raw_tissue_reading

    def get_raw_reference_reading(self) -> NDArray[(2, Any), float]:
        if self._raw_ref_reading is None:
            self._raw_ref_reading = np.zeros(self.get_raw_tissue_reading().shape)
        return self._raw_ref_reading
