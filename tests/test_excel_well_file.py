# -*- coding: utf-8 -*-

import datetime
import os
from shutil import copy
import tempfile

from curibio.sdk import ExcelWellFile
from curibio.sdk import MetadataNotFoundError
from curibio.sdk import PlateRecording
from mantarray_file_manager import CURI_BIO_ACCOUNT_UUID
from mantarray_file_manager import CURI_BIO_USER_ACCOUNT_ID
from mantarray_file_manager import METADATA_UUID_DESCRIPTIONS
from mantarray_file_manager import WELL_NAME_UUID
import numpy as np
from openpyxl import load_workbook
import pytest
from stdlib_utils import get_current_file_abs_directory

from .fixtures import fixture_generic_excel_well_file_0_1_0
from .utils import get_cell_value


__fixtures__ = [
    fixture_generic_excel_well_file_0_1_0,
]

PATH_OF_CURRENT_FILE = get_current_file_abs_directory()


def test_ExcelWellFile__opens_and_get_file_name():
    file_name = os.path.join("excel_optical_data", "optical_data_filled_template.xlsx")
    wf = ExcelWellFile(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            file_name,
        )
    )
    expected_path = os.path.join(PATH_OF_CURRENT_FILE, file_name)
    assert wf.get_file_name() == expected_path


def test_ExcelWellFile__get_excel_metadata_value__raises_error_when_metadata_is_missing():
    wf = ExcelWellFile(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            "excel_optical_data",
            "optical_data_missing_well_name.xlsx",
        )
    )
    test_metadata_uuid = WELL_NAME_UUID
    with pytest.raises(
        MetadataNotFoundError, match=METADATA_UUID_DESCRIPTIONS[test_metadata_uuid]
    ):
        wf.get_excel_metadata_value(test_metadata_uuid)


def test_ExcelWellFile__get_file_version(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_file_version() == "0.1.1"


def test_ExcelWellFile__get_unique_recording_key(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_unique_recording_key() == (
        "Test Barcode",
        datetime.datetime(2020, 10, 12, 4, 40),
    )


def test_ExcelWellFile__get_well_name(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_well_name() == "A1"


def test_ExcelWellFile__get_well_index(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_well_index() == 0


def test_ExcelWellFile__get_plate_barcode(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_plate_barcode() == "Test Barcode"


def test_ExcelWellFile__get_user_account(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_user_account() == CURI_BIO_USER_ACCOUNT_ID


def test_ExcelWellFile__get_timestamp_of_beginning_of_data_acquisition(
    generic_excel_well_file_0_1_0,
):
    assert (
        generic_excel_well_file_0_1_0.get_timestamp_of_beginning_of_data_acquisition()
        == datetime.datetime(2020, 10, 12, 4, 40)
    )


def test_ExcelWellFile__get_customer_account(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_customer_account() == CURI_BIO_ACCOUNT_UUID


def test_ExcelWellFile__get_mantarray_serial_number(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_mantarray_serial_number() == "Test Name"


def test_ExcelWellFile__get_begin_recording(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_begin_recording() == datetime.datetime(
        2020, 10, 12, 4, 40
    )


def test_ExcelWellFile__get_timestamp_of_first_tissue_data_point(
    generic_excel_well_file_0_1_0,
):
    assert (
        generic_excel_well_file_0_1_0.get_timestamp_of_first_tissue_data_point()
        == datetime.datetime(2020, 10, 12, 4, 40)
    )


def test_ExcelWellFile__get_timestamp_of_first_ref_data_point(
    generic_excel_well_file_0_1_0,
):
    assert (
        generic_excel_well_file_0_1_0.get_timestamp_of_first_ref_data_point()
        == datetime.datetime(2020, 10, 12, 4, 40)
    )


def test_ExcelWellFile__get_tissue_sampling_period_microseconds(
    generic_excel_well_file_0_1_0,
):
    assert (
        generic_excel_well_file_0_1_0.get_tissue_sampling_period_microseconds()
        == 0.016666e6
    )


def test_ExcelWellFile__get_reference_sampling_period_microseconds(
    generic_excel_well_file_0_1_0,
):
    assert (
        generic_excel_well_file_0_1_0.get_reference_sampling_period_microseconds() == 0
    )


def test_ExcelWellFile__get_recording_start_index(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_recording_start_index() == 0


def test_ExcelWellFile__get_twitches_point_up(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_twitches_point_up() is True


def test_ExcelWellFile__get_raw_tissue_reading(generic_excel_well_file_0_1_0):
    raw_tissue_reading = generic_excel_well_file_0_1_0.get_raw_tissue_reading()
    assert raw_tissue_reading.shape == (2, 1466)
    np.testing.assert_almost_equal(raw_tissue_reading[0][0], 0.066733333)
    np.testing.assert_almost_equal(raw_tissue_reading[0][1], 0.083416667)
    np.testing.assert_almost_equal(raw_tissue_reading[0][-1], 24.50781667)
    np.testing.assert_almost_equal(raw_tissue_reading[1][0], -0.007479378)
    np.testing.assert_almost_equal(raw_tissue_reading[1][-1], 104.7408281)


def test_ExcelWellFile__get_raw_reference_reading(generic_excel_well_file_0_1_0):
    assert generic_excel_well_file_0_1_0.get_raw_reference_reading().shape == (2, 1466)


def test_ExcelWellFile__get_interpolation_value__returns_data_period_if_no_value_is_given():
    wf = ExcelWellFile(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            "excel_optical_data",
            "optical_data_missing_well_name.xlsx",
        )
    )
    assert wf.get_interpolation_value() == 0.016666e6


def test_ExcelWellFile__get_interpolation_value__returns_correct_value_when_given(
    generic_excel_well_file_0_1_0,
):
    assert generic_excel_well_file_0_1_0.get_interpolation_value() == 0.0123e6


def test_PlateRecording__creates_a_pipeline_with_no_filter_and_correct_sampling_period(
    generic_excel_well_file_0_1_0,
):
    pr = PlateRecording([generic_excel_well_file_0_1_0])
    actual = pr.get_pipeline_template()
    assert actual.noise_filter_uuid is None
    assert actual.tissue_sampling_period == 1666.6


def test_PlateRecording_write_xlsx__creates_continuous_recording_sheet__with_single_optical_well_data(
    generic_excel_well_file_0_1_0,
):
    pr = PlateRecording([generic_excel_well_file_0_1_0])

    expected_file_name = "test_optical_file.xlsx"
    with tempfile.TemporaryDirectory() as tmp_dir:
        pr.write_xlsx(
            tmp_dir, file_name=expected_file_name, create_waveform_charts=False
        )
        actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
        actual_sheet = actual_workbook[actual_workbook.sheetnames[1]]

        assert get_cell_value(actual_sheet, 0, 0) == "Time (seconds)"
        assert get_cell_value(actual_sheet, 1, 0) == 0.0123
        assert get_cell_value(actual_sheet, 10, 0) == 0.0123 * 10

        assert get_cell_value(actual_sheet, 0, 1) == "A1"
        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 1, 1), 0.778961602926782, 6
        )
        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 10, 1), 0.1535264502266848, 6
        )


def test_PlateRecording_write_xlsx__creates_continuous_recording_sheet__with_correct_peaks_and_valleys_for_optical_data(
    generic_excel_well_file_0_1_0,
):
    pr = PlateRecording([generic_excel_well_file_0_1_0])

    expected_file_name = "test_optical_file.xlsx"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir = "."
        pr.write_xlsx(
            tmp_dir, file_name=expected_file_name, create_waveform_charts=False
        )
        actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
        actual_sheet = actual_workbook[actual_workbook.sheetnames[1]]

        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 47, 101), 293.781693313573, 6
        )
        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 1938, 101), 390.9615827, 6
        )
        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 115, 100), -0.792794065, 6
        )
        np.testing.assert_almost_equal(
            get_cell_value(actual_sheet, 1882, 100), 100.2838749, 6
        )


def test_PlateRecording__can_be_initialized_from_zipped_optical_files():
    file_name = "zipped_optical_files.zip"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(PATH_OF_CURRENT_FILE, "excel_optical_data", file_name),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == (0, 4)
        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open


def test_PlateRecording__can_be_initialized_from_zipped_optical_file_folder():
    file_name = "zipped_optical_file_folder.zip"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(PATH_OF_CURRENT_FILE, "excel_optical_data", file_name),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == (0, 4)
        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open
