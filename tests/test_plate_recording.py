# -*- coding: utf-8 -*-
"""Tests for PlateRecording subclass.

To create a file to look at: python3 -c "import os; from curibio.sdk import PlateRecording; PlateRecording([os.path.join('tests','h5','v0.3.1','MA20123456__2020_08_17_145752__A1.h5')]).write_xlsx('.',file_name='temp.xlsx')"
To create a file to look at: python3 -c "import os; from curibio.sdk import PlateRecording; PlateRecording([os.path.join('tests','h5','v0.3.1','MA201110001__2020_09_03_213024__A3.h5')]).write_xlsx('.',file_name='temp.xlsx')"
To create a file to look at: python3 -c "import os; from curibio.sdk import PlateRecording; PlateRecording.from_directory(os.path.join('tests','h5','v0.3.1')).write_xlsx('.',file_name='temp.xlsx')"


"""
import datetime
import os
from shutil import copy
import tempfile
from typing import Optional
from typing import Union

from curibio.sdk import __version__
from curibio.sdk import AGGREGATE_METRICS_SHEET_NAME
from curibio.sdk import CALCULATED_METRIC_DISPLAY_NAMES
from curibio.sdk import CONTINUOUS_WAVEFORM_SHEET_NAME
from curibio.sdk import INTERPOLATED_DATA_PERIOD_CMS
from curibio.sdk import METADATA_EXCEL_SHEET_NAME
from curibio.sdk import METADATA_INSTRUMENT_ROW_START
from curibio.sdk import METADATA_OUTPUT_FILE_ROW_START
from curibio.sdk import METADATA_RECORDING_ROW_START
from curibio.sdk import NUMBER_OF_PER_TWITCH_METRICS
from curibio.sdk import PER_TWITCH_METRICS_SHEET_NAME
from curibio.sdk import plate_recording
from curibio.sdk import PlateRecording
from freezegun import freeze_time
from labware_domain_models import LabwareDefinition
from mantarray_file_manager import MAIN_FIRMWARE_VERSION_UUID
from mantarray_file_manager import MANTARRAY_SERIAL_NUMBER_UUID
from mantarray_file_manager import METADATA_UUID_DESCRIPTIONS
from mantarray_file_manager import PLATE_BARCODE_UUID
from mantarray_file_manager import SOFTWARE_BUILD_NUMBER_UUID
from mantarray_file_manager import SOFTWARE_RELEASE_VERSION_UUID
from mantarray_file_manager import UTC_BEGINNING_RECORDING_UUID
from mantarray_waveform_analysis import BESSEL_LOWPASS_10_UUID
from mantarray_waveform_analysis import BUTTERWORTH_LOWPASS_30_UUID
from mantarray_waveform_analysis import CENTIMILLISECONDS_PER_SECOND
from mantarray_waveform_analysis import Pipeline
from mantarray_waveform_analysis import TooFewPeaksDetectedError
from mantarray_waveform_analysis import TwoPeaksInARowError
from mantarray_waveform_analysis import TwoValleysInARowError
from matplotlib.figure import Figure
from openpyxl import load_workbook
from PIL import Image
from PIL import ImageDraw
from PIL.PngImagePlugin import PngImageFile
import pytest
from pytest import approx
from stdlib_utils import get_current_file_abs_directory
from stdlib_utils import is_system_windows

from .fixtures import fixture_generic_well_file_0_3_1
from .fixtures import fixture_generic_well_file_0_3_1__2
from .fixtures import fixture_generic_well_file_0_3_2
from .fixtures import fixture_plate_recording_in_tmp_dir_for_24_wells_0_3_2
from .fixtures import fixture_plate_recording_in_tmp_dir_for_generic_well_file_0_3_1
from .fixtures import fixture_plate_recording_in_tmp_dir_for_generic_well_file_0_3_2
from .fixtures import fixture_plate_recording_in_tmp_dir_for_multiple_well_files_0_3_1
from .fixtures import fixture_plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1
from .fixtures import fixture_real_3min_well_file_0_3_1
from .utils import get_cell_value

__fixtures__ = (
    fixture_generic_well_file_0_3_1,
    fixture_generic_well_file_0_3_2,
    fixture_plate_recording_in_tmp_dir_for_generic_well_file_0_3_2,
    fixture_plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
    fixture_plate_recording_in_tmp_dir_for_multiple_well_files_0_3_1,
    fixture_generic_well_file_0_3_1__2,
    fixture_real_3min_well_file_0_3_1,
    fixture_plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1,
    fixture_plate_recording_in_tmp_dir_for_24_wells_0_3_2,
)

PATH_OF_CURRENT_FILE = get_current_file_abs_directory()


def test_init__creates_a_pipeline_template_with_correct_sampling_frequency_and_bessel_for_960cms_if_none_is_given(
    generic_well_file_0_3_1,
):
    pr = PlateRecording([generic_well_file_0_3_1])
    actual = pr.get_pipeline_template()
    assert actual.noise_filter_uuid == BESSEL_LOWPASS_10_UUID
    assert actual.tissue_sampling_period == 960


def test_init__creates_a_pipeline_template_with_correct_sampling_frequency_and_butterworth_for_160cms_if_none_is_given(
    generic_well_file_0_3_2,
):
    pr = PlateRecording([generic_well_file_0_3_2])
    actual = pr.get_pipeline_template()
    assert actual.noise_filter_uuid == BUTTERWORTH_LOWPASS_30_UUID
    assert actual.tissue_sampling_period == 160


def test_write_xlsx__creates_file_at_supplied_path_and_name(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1
    file_dir = tmp_dir
    file_name = "my_file.xlsx"
    pr.write_xlsx(file_dir, file_name=file_name)
    assert os.path.exists(os.path.join(file_dir, file_name)) is True


@pytest.mark.slow
def test_write_xlsx__creates_file_for_all_24_wells(
    plate_recording_in_tmp_dir_for_24_wells_0_3_2,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_24_wells_0_3_2
    file_name = "my_file2.xlsx"
    pr.write_xlsx(tmp_dir, file_name=file_name)
    assert os.path.exists(os.path.join(tmp_dir, file_name)) is True


def test_write_xlsx__creates_file_at_supplied_path_with_auto_generated_name(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1

    pr.write_xlsx(tmp_dir)
    expected_file_name = "MA20123456__2020_08_17_145810.xlsx"
    assert os.path.exists(os.path.join(tmp_dir, expected_file_name)) is True


def test_write_xlsx__creates_per_twitch_metrics_sheet_labels(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_2,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_2

    pr.write_xlsx(tmp_dir, create_waveform_charts=False)
    expected_file_name = "MA20223322__2020_09_02_173943.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[5] == PER_TWITCH_METRICS_SHEET_NAME
    curr_sheet = actual_workbook[PER_TWITCH_METRICS_SHEET_NAME]
    curr_row = 0
    assert get_cell_value(curr_sheet, curr_row, 0) == "A1"
    assert get_cell_value(curr_sheet, curr_row, 1) is None

    curr_row += 1

    assert get_cell_value(curr_sheet, curr_row, 0) == "Timepoint of Twitch Contraction"
    curr_row += 1
    assert get_cell_value(curr_sheet, curr_row, 0) == "Twitch Period (seconds)"
    curr_row += 1
    assert get_cell_value(curr_sheet, curr_row, 0) == "Twitch Frequency (Hz)"
    curr_row += 1
    assert get_cell_value(curr_sheet, curr_row, 0) == "Twitch Amplitude"
    curr_row += 1
    assert get_cell_value(curr_sheet, curr_row, 0) == "Twitch Width 50 (FWHM) (seconds)"
    curr_row += 1

    curr_row += (
        NUMBER_OF_PER_TWITCH_METRICS - 5
    )  # subtract the amount of the metrics that we already wrote assert statements for and increment the curr_row
    curr_row += 1  # gap between data for the different wells
    assert get_cell_value(curr_sheet, curr_row, 0) == "B1"
    assert get_cell_value(curr_sheet, curr_row, 1) is None
    curr_row += (NUMBER_OF_PER_TWITCH_METRICS + 2) * 3  # A2 row number
    assert get_cell_value(curr_sheet, curr_row, 0) == "A2"
    assert get_cell_value(curr_sheet, curr_row, 1) == "Twitch 1"
    assert get_cell_value(curr_sheet, curr_row, 11) == "Twitch 11"


def test_write_xlsx__writes_in_per_twitch_metrics_sheet_for_single_well(
    plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1, real_3min_well_file_0_3_1
):

    pr, tmp_dir = plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1

    pr.write_xlsx(tmp_dir, create_continuous_waveforms=False)
    expected_file_name = "MA201110001__2020_09_03_213044.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[5] == PER_TWITCH_METRICS_SHEET_NAME
    curr_sheet = actual_workbook[PER_TWITCH_METRICS_SHEET_NAME]
    curr_row = 0
    curr_row += 8 * (NUMBER_OF_PER_TWITCH_METRICS + 2)
    expected_number_twitches = 429
    assert (
        get_cell_value(curr_sheet, curr_row, expected_number_twitches)
        == f"Twitch {expected_number_twitches}"
    )
    curr_row += 1
    assert ("timepoint", get_cell_value(curr_sheet, curr_row, 1)) == (
        "timepoint",
        66185 / CENTIMILLISECONDS_PER_SECOND,
    )
    curr_row += 1
    assert ("period", get_cell_value(curr_sheet, curr_row, 1)) == (
        "period",
        50880 / CENTIMILLISECONDS_PER_SECOND,
    )
    assert (
        "period",
        get_cell_value(curr_sheet, curr_row, expected_number_twitches),
    ) == (
        "period",
        50880 / CENTIMILLISECONDS_PER_SECOND,
    )
    curr_row += 1
    curr_row += 1
    assert ("amplitude", get_cell_value(curr_sheet, curr_row, 1)) == (
        "amplitude",
        84937,
    )
    assert (
        "amplitude",
        get_cell_value(curr_sheet, curr_row, expected_number_twitches),
    ) == (
        "amplitude",
        104234,
    )
    curr_row += 1
    assert ("twitch width 50", get_cell_value(curr_sheet, curr_row, 1)) == (
        "twitch width 50",
        0.25007,
    )
    assert (
        "twitch width 50",
        get_cell_value(curr_sheet, curr_row, expected_number_twitches),
    ) == ("twitch width 50", 0.25806)


def test_write_xlsx__creates_aggregate_metrics_sheet_labels(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_2,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_2

    pr.write_xlsx(tmp_dir, create_waveform_charts=False)
    expected_file_name = "MA20223322__2020_09_02_173943.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[4] == AGGREGATE_METRICS_SHEET_NAME
    aggregate_metrics_sheet = actual_workbook[AGGREGATE_METRICS_SHEET_NAME]
    curr_row = 0
    assert get_cell_value(aggregate_metrics_sheet, curr_row, 2) == "A1"
    curr_row += 1
    assert (
        get_cell_value(aggregate_metrics_sheet, curr_row, 1) == "Treatment Description"
    )
    curr_row += 1
    assert get_cell_value(aggregate_metrics_sheet, curr_row, 1) == "n (twitches)"

    curr_row += 1
    # check the labels in columns A & B
    for iter_idx, (_, iter_metric_name) in enumerate(
        CALCULATED_METRIC_DISPLAY_NAMES.items()
    ):
        curr_row += 1
        actual_metric_name = get_cell_value(aggregate_metrics_sheet, curr_row, 0)
        if isinstance(iter_metric_name, tuple):
            _, iter_metric_name = iter_metric_name

        assert (iter_idx, actual_metric_name) == (iter_idx, iter_metric_name)
        for iter_sub_metric_idx, iter_sub_metric_name in enumerate(
            ("Mean", "StDev", "CoV", "SEM")
        ):
            actual_sub_metric_name = get_cell_value(
                aggregate_metrics_sheet, curr_row, 1
            )
            curr_row += 1
            assert (iter_idx, iter_sub_metric_idx, actual_sub_metric_name) == (
                iter_idx,
                iter_sub_metric_idx,
                iter_sub_metric_name,
            )


def test_write_xlsx__writes_in_aggregate_metrics_for_single_well(
    plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1, real_3min_well_file_0_3_1
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_real_3min_well_file_0_3_1

    pr.write_xlsx(tmp_dir, create_continuous_waveforms=False)
    expected_file_name = "MA201110001__2020_09_03_213044.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[4] == AGGREGATE_METRICS_SHEET_NAME
    actual_sheet = actual_workbook[AGGREGATE_METRICS_SHEET_NAME]
    well_idx = real_3min_well_file_0_3_1.get_well_index()
    curr_row = 2
    actual_num_twitches = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_num_twitches == 429

    # Period
    curr_row += 2
    actual_mean = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_mean == 0.51272
    curr_row += 1
    actual_stdev = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_stdev == 0.00472
    curr_row += 1
    actual_cov = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_cov == approx(0.00472 / 0.51272)
    curr_row += 1
    actual_sem = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_sem == approx(0.00472 / 429 ** 0.5)

    # Twitch Width 50
    curr_row += 2 + 5 + 5
    expected_stdev = 0.00254
    expected_mean = 0.25524
    actual_mean = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_mean == approx(expected_mean)
    curr_row += 1
    actual_stdev = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_stdev == approx(expected_stdev)
    curr_row += 1
    actual_cov = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_cov == approx(expected_stdev / expected_mean)
    curr_row += 1
    actual_sem = get_cell_value(actual_sheet, curr_row, 2 + well_idx)
    assert actual_sem == approx(expected_stdev / 429 ** 0.5)


def test_write_xlsx__creates_metadata_sheet_with_recording_info(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1
    file_dir = tmp_dir

    pr.write_xlsx(file_dir, create_waveform_charts=False)
    expected_file_name = "MA20123456__2020_08_17_145810.xlsx"
    actual_workbook = load_workbook(os.path.join(file_dir, expected_file_name))
    assert actual_workbook.sheetnames[0] == METADATA_EXCEL_SHEET_NAME

    metadata_sheet = actual_workbook[METADATA_EXCEL_SHEET_NAME]
    assert (
        get_cell_value(metadata_sheet, METADATA_RECORDING_ROW_START, 0)
        == "Recording Information:"
    )
    for iter_row, metadata_uuid, expected_value in [
        (0, PLATE_BARCODE_UUID, "MA20123456"),
        (
            1,
            UTC_BEGINNING_RECORDING_UUID,
            datetime.datetime(2020, 8, 17, 14, 58, 10, 728253),
        ),
    ]:
        actual_label = get_cell_value(
            metadata_sheet, METADATA_RECORDING_ROW_START + 1 + iter_row, 1
        )
        actual_value = get_cell_value(
            metadata_sheet, METADATA_RECORDING_ROW_START + 1 + iter_row, 2
        )

        assert (iter_row, actual_label) == (
            iter_row,
            METADATA_UUID_DESCRIPTIONS[metadata_uuid],
        )
        assert (iter_row, actual_value) == (iter_row, expected_value)


def test_write_xlsx__creates_metadata_sheet_with_mantarray_info(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1

    pr.write_xlsx(tmp_dir, create_waveform_charts=False)
    expected_file_name = "MA20123456__2020_08_17_145810.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[0] == METADATA_EXCEL_SHEET_NAME

    metadata_sheet = actual_workbook[METADATA_EXCEL_SHEET_NAME]
    curr_row = METADATA_INSTRUMENT_ROW_START
    assert get_cell_value(metadata_sheet, curr_row, 0) == "Device Information:"
    curr_row += 1
    assert get_cell_value(metadata_sheet, curr_row, 1) == "H5 File Layout Version"
    assert get_cell_value(metadata_sheet, curr_row, 2) == "0.3.1"
    curr_row += 1
    for iter_row, metadata_uuid, expected_value in [
        (0, MANTARRAY_SERIAL_NUMBER_UUID, "M02001900"),
        (1, SOFTWARE_RELEASE_VERSION_UUID, "0.2.2"),
        (2, SOFTWARE_BUILD_NUMBER_UUID, "200817143923--820"),
        (3, MAIN_FIRMWARE_VERSION_UUID, "0.0.0"),
    ]:
        actual_label = get_cell_value(metadata_sheet, curr_row + iter_row, 1)
        actual_value = get_cell_value(metadata_sheet, curr_row + iter_row, 2)

        assert (iter_row, actual_label) == (
            iter_row,
            METADATA_UUID_DESCRIPTIONS[metadata_uuid],
        )
        assert (iter_row, actual_value) == (iter_row, expected_value)


@freeze_time("2020-09-02 12:20:12.223344")
def test_write_xlsx__creates_metadata_sheet_with_output_format_info(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1

    pr.write_xlsx(tmp_dir, create_waveform_charts=False)
    expected_file_name = "MA20123456__2020_08_17_145810.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    assert actual_workbook.sheetnames[0] == METADATA_EXCEL_SHEET_NAME

    metadata_sheet = actual_workbook[METADATA_EXCEL_SHEET_NAME]
    curr_row = METADATA_OUTPUT_FILE_ROW_START
    assert get_cell_value(metadata_sheet, curr_row, 0) == "Output Format:"

    curr_row += 1
    assert get_cell_value(metadata_sheet, curr_row, 1) == "SDK Version"
    assert get_cell_value(metadata_sheet, curr_row, 2) == __version__

    curr_row += 1
    assert get_cell_value(metadata_sheet, curr_row, 1) == "File Creation Timestamp"
    assert get_cell_value(metadata_sheet, curr_row, 2) == datetime.datetime(
        2020, 9, 2, 12, 20, 12
    )


def test_write_xlsx__creates_continuous_recording_sheet__with_multiple_well_data(
    plate_recording_in_tmp_dir_for_multiple_well_files_0_3_1,
):
    pr, tmp_dir = plate_recording_in_tmp_dir_for_multiple_well_files_0_3_1

    pr.write_xlsx(tmp_dir, create_waveform_charts=False)
    expected_file_name = "MA20123456__2020_08_17_145810.xlsx"
    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    expected_sheet_name = CONTINUOUS_WAVEFORM_SHEET_NAME
    assert actual_workbook.sheetnames[1] == expected_sheet_name
    actual_sheet = actual_workbook[expected_sheet_name]

    assert actual_sheet.cell(row=0 + 1, column=1 + 1).value == "A1"
    assert actual_sheet.cell(row=0 + 1, column=24 + 1).value == "D6"

    assert actual_sheet.cell(row=0 + 1, column=0 + 1).value == "Time (seconds)"
    assert (
        actual_sheet.cell(row=1 + 1, column=0 + 1).value
        == INTERPOLATED_DATA_PERIOD_CMS / CENTIMILLISECONDS_PER_SECOND
    )
    assert (
        actual_sheet.cell(row=10 + 1, column=0 + 1).value
        == 10 * INTERPOLATED_DATA_PERIOD_CMS / CENTIMILLISECONDS_PER_SECOND
    )

    assert get_cell_value(actual_sheet, 0, 5) == "A2"
    assert get_cell_value(actual_sheet, 1, 10) == -1238675
    assert get_cell_value(actual_sheet, 10, 10) == -1531018.5

    assert get_cell_value(actual_sheet, 0, 10) == "B3"
    assert actual_sheet.cell(row=1 + 1, column=10 + 1).value == -1238675
    assert actual_sheet.cell(row=10 + 1, column=10 + 1).value == -1531018.5


def test_PlateRecording__init_pipelines__does_not_init_twice(
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
):
    pr, _ = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1
    pr._init_pipelines()  # pylint:disable=protected-access # Eli (9/10/20): this is a performance optimization, so this is just an internal implementation detail
    initial_pipelines = (
        pr._pipelines  # pylint:disable=protected-access # Eli (9/10/20): this is a performance optimization, so this is just an internal implementation detail
    )
    pr._init_pipelines()  # pylint:disable=protected-access # Eli (9/10/20): this is a performance optimization, so this is just an internal implementation detail
    second_call_pipelines = (
        pr._pipelines  # pylint:disable=protected-access # Eli (9/10/20): this is a performance optimization, so this is just an internal implementation detail
    )
    assert second_call_pipelines is initial_pipelines


twenty_four_well = LabwareDefinition(row_count=4, column_count=6)


def process_region(
    image: PngImageFile, x: int, y: int, width: int, height: int
) -> Optional[Union[int, float]]:
    # pylint:disable=inconsistent-return-statements,bare-except
    # Eli (9/11/20): this ability to do visual regression testing is evolving. If it works then we will recreate it properly from the template code in the Medium article
    region_total: Union[float, int] = 0
    # This can be used as the sensitivity factor, the larger it is the less sensitive the comparison
    factor = 100

    for coordinateY in range(y, y + height):
        for coordinateX in range(x, x + width):
            try:
                pixel = image.getpixel((coordinateX, coordinateY))
                region_total += sum(pixel) / 4
            except:  # noqa: E722 # Eli (9/11/20): this ability to do visual regression testing is evolving. If it works then we will recreate it properly from the template code in the Medium article
                return None

    return region_total / factor


def matplotlib_visual_regression(
    png_file_path_and_prefix: str, the_figure: Figure
) -> None:
    # based on https://blog.rinatussenov.com/automating-manual-visual-regression-tests-with-python-and-selenium-be66be950196
    base_file_path = png_file_path_and_prefix + "-base.png"
    if not os.path.exists(base_file_path):
        print(  # allow-print
            f"No base image detected for {base_file_path} so saving a new one."
        )
        the_figure.savefig(base_file_path)
        return
    actual_file_path = png_file_path_and_prefix + "-actual.png"
    the_figure.savefig(actual_file_path)

    screenshot_staging = Image.open(actual_file_path)
    screenshot_production = Image.open(base_file_path)
    columns = 60
    rows = 80
    screen_width, screen_height = screenshot_staging.size

    block_width = ((screen_width - 1) // columns) + 1  # this is just a division ceiling
    block_height = ((screen_height - 1) // rows) + 1
    problem_found = False
    for y in range(0, screen_height, block_height + 1):
        for x in range(0, screen_width, block_width + 1):
            region_staging = process_region(
                screenshot_staging, x, y, block_width, block_height
            )
            region_production = process_region(
                screenshot_production, x, y, block_width, block_height
            )

            if (
                region_staging is not None
                and region_production is not None
                and region_production != region_staging
            ):
                problem_found = True
                draw = ImageDraw.Draw(screenshot_staging)
                draw.rectangle((x, y, x + block_width, y + block_height), outline="red")
    if problem_found:
        screenshot_staging.save(png_file_path_and_prefix + "-diff.png")
        raise Exception(f"Visual Mismatch for {png_file_path_and_prefix}")


@pytest.mark.slow
def test_PlateRecording__create_stacked_plot_for_24_wells():
    pr = PlateRecording.from_directory(
        os.path.join(
            PATH_OF_CURRENT_FILE, "h5", "v0.3.1", "MA201110001__2020_09_03_213024"
        )
    )
    fig = pr.create_stacked_plot()
    if not is_system_windows():
        matplotlib_visual_regression(
            os.path.join(
                PATH_OF_CURRENT_FILE, "visual-regression-testing", "stacked-24"
            ),
            fig,
        )


@pytest.mark.parametrize(
    "expected_error,error_message,test_description",
    [
        (
            TwoPeaksInARowError(([], []), [], (0, 1)),
            "Error: Two Contractions in a Row Detected",
            "handles TwoPeaksInARowError",
        ),
        (
            TwoValleysInARowError(([], []), [], (0, 1)),
            "Error: Two Relaxations in a Row Detected",
            "handles TwoValleysInARowError",
        ),
        (
            TooFewPeaksDetectedError(),
            "Error: Not Enough Twitches Detected",
            "handles TooFewPeaksDetectedError",
        ),
    ],
)
def test_write_xlsx__writes_NA_if_peak_detections_errors_in_aggregate_metrics(
    mocker,
    plate_recording_in_tmp_dir_for_generic_well_file_0_3_1,
    expected_error,
    error_message,
    test_description,
):
    error_val = "N/A"
    pr, tmp_dir = plate_recording_in_tmp_dir_for_generic_well_file_0_3_1
    expected_file_name = "temp.xlsx"
    mocker.patch.object(
        Pipeline, "get_magnetic_data_metrics", autospec=True, side_effect=expected_error
    )
    pr.write_xlsx(tmp_dir, file_name=expected_file_name, create_waveform_charts=False)

    actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))
    aggregate_metrics_sheet = actual_workbook[AGGREGATE_METRICS_SHEET_NAME]

    expected_well_idx = 11
    curr_row = 0
    assert get_cell_value(aggregate_metrics_sheet, curr_row, expected_well_idx) == "B3"
    curr_row += 2
    assert get_cell_value(aggregate_metrics_sheet, curr_row, 1) == "n (twitches)"
    assert (
        get_cell_value(aggregate_metrics_sheet, curr_row, expected_well_idx)
        == error_val
    )
    curr_row += 1
    assert (
        get_cell_value(aggregate_metrics_sheet, curr_row, expected_well_idx)
        == error_message
    )

    curr_row += 1
    for (
        _,
        iter_metric_name,
    ) in CALCULATED_METRIC_DISPLAY_NAMES.items():
        for iter_sub_metric_name in ("Mean", "StDev", "CoV", "SEM"):
            actual_sub_metric_name = get_cell_value(
                aggregate_metrics_sheet, curr_row, 1
            )
            assert (iter_metric_name, iter_sub_metric_name, actual_sub_metric_name) == (
                iter_metric_name,
                iter_sub_metric_name,
                iter_sub_metric_name,
            )
            actual_sub_metric_val = get_cell_value(
                aggregate_metrics_sheet, curr_row, expected_well_idx
            )
            assert (iter_metric_name, iter_sub_metric_name, actual_sub_metric_val) == (
                iter_metric_name,
                iter_sub_metric_name,
                error_val,
            )
            curr_row += 1
        assert (
            iter_metric_name is not None
            and get_cell_value(aggregate_metrics_sheet, curr_row, expected_well_idx)
            is None
        )
        curr_row += 1


@pytest.mark.slow
def test_PlateRecording__write_xlsx__logs_progress(mocker):
    spied_info_logger = mocker.spy(plate_recording.logger, "info")

    pr = PlateRecording.from_directory(
        os.path.join(
            PATH_OF_CURRENT_FILE, "h5", "v0.3.1", "MA201110001__2020_09_03_213024"
        )
    )
    with tempfile.TemporaryDirectory() as tmp_dir:
        pr.write_xlsx(tmp_dir)

    spied_info_logger.assert_any_call("Loading data from H5 file(s)")
    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 0% (Well A1, 1 out of 24)"
    )
    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 17% (Well A2, 5 out of 24)"
    )
    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 96% (Well D6, 24 out of 24)"
    )
    spied_info_logger.assert_any_call("Opening .xlsx file")
    spied_info_logger.assert_any_call("Writing H5 file metadata")
    spied_info_logger.assert_any_call("Creating waveform data sheet")
    spied_info_logger.assert_any_call("Writing waveform data of well A1 (1 out of 24)")
    spied_info_logger.assert_any_call("Creating chart of waveform data of well A1")
    spied_info_logger.assert_any_call(
        "Adding peak and valley markers to chart of well A1"
    )
    spied_info_logger.assert_any_call("Writing waveform data of well D6 (24 out of 24)")
    spied_info_logger.assert_any_call("Creating chart of waveform data of well D6")
    spied_info_logger.assert_any_call(
        "Adding peak and valley markers to chart of well D6"
    )
    spied_info_logger.assert_any_call("Creating aggregate metrics sheet")
    for (_, metric) in CALCULATED_METRIC_DISPLAY_NAMES.items():
        if isinstance(metric, tuple):
            _, metric = metric
        for submetric in ("Mean", "StDev", "CoV", "SEM"):
            spied_info_logger.assert_any_call(f"Writing {submetric} of {metric}")
            spied_info_logger.assert_any_call(f"Writing {submetric} of {metric}")
            spied_info_logger.assert_any_call(f"Writing {submetric} of {metric}")
    spied_info_logger.assert_any_call("Saving .xlsx file")
    spied_info_logger.assert_any_call("Done writing to .xlsx")


def test_PlateRecording__can_write_file_of_v0_1_1_to_xlsx():
    pr = PlateRecording.from_directory(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            "h5",
            "v0.1.1",
        )
    )

    with tempfile.TemporaryDirectory() as tmp_dir:
        pr.write_xlsx(tmp_dir, create_waveform_charts=False)

        expected_file_name = "MA20001123__2020_08_20_170600.xlsx"
        actual_workbook = load_workbook(os.path.join(tmp_dir, expected_file_name))

        metadata_sheet = actual_workbook[METADATA_EXCEL_SHEET_NAME]
        assert metadata_sheet.cell(row=5 + 1, column=2 + 1).value == "0.1.1"

        waveform_sheet = actual_workbook[CONTINUOUS_WAVEFORM_SHEET_NAME]
        assert waveform_sheet.cell(row=0 + 1, column=1 + 1).value == "A1"
        assert waveform_sheet.cell(row=0 + 1, column=0 + 1).value == "Time (seconds)"
        assert (
            waveform_sheet.cell(row=1 + 1, column=0 + 1).value
            == INTERPOLATED_DATA_PERIOD_CMS / CENTIMILLISECONDS_PER_SECOND
        )


def test_PlateRecording__get_reference_magnetic_data__returns_expected_values(
    generic_well_file_0_3_2,
):
    pr = PlateRecording([generic_well_file_0_3_2])
    actual = pr.get_reference_magnetic_data(4)
    assert actual.shape[0] == 2
    assert actual[1][0] == -376649
    assert actual[1][-1] == -552810


def test_PlateRecording__get_reference_magnetic_data__returns_expected_values_with_file_version_0_1_1():
    pr = PlateRecording.from_directory(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            "h5",
            "v0.1.1",
        )
    )
    actual = pr.get_reference_magnetic_data(0)
    assert actual.shape[0] == 2
    assert actual[1][0] == 78371
    assert actual[1][-1] == 116599


def test_PlateRecording__can_be_initialized_from_zipped_files__and_logs_process_correctly(
    mocker,
):
    spied_info_logger = mocker.spy(plate_recording.logger, "info")

    file_name = "MA20123456__2020_08_17_145752_files.zip"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(
                PATH_OF_CURRENT_FILE, "zipped_MA20123456__2020_08_17_145752", file_name
            ),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == (0, 4, 8)

        pr.write_xlsx(tmp_dir)

        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open

    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 0% (Well A1, 1 out of 3)"
    )
    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 33% (Well A2, 2 out of 3)"
    )
    spied_info_logger.assert_any_call(
        "Loading tissue and reference data... 67% (Well A3, 3 out of 3)"
    )


def test_PlateRecording__can_be_initialized_from_a_mac_zipped_folder():
    file_name = "MA20123456__2020_08_17_145752_folder.zip"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(PATH_OF_CURRENT_FILE, "zipped_mac_folder", file_name),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == (0, 4, 8)
        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open


def test_PlateRecording__can_be_initialized_from_a_windows_zipped_folder():
    file_name = "MA 20 PEI Test Plate 2 - 2020_08_13_153638-20201007T210515Z-001.zip"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(PATH_OF_CURRENT_FILE, "zipped_windows_folder", file_name),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == tuple(range(24))

        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open


@pytest.mark.slow
def test_PlateRecording__can_be_initialized_from_data_having_iterpolation_issue():
    file_name = "MA 22 Plate 3_2020_08_25_164108.zip"
    expected_excel_file = "test_file"
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file_path = os.path.join(tmp_dir, file_name)
        copy(
            os.path.join(PATH_OF_CURRENT_FILE, "h5", "interpolation_error", file_name),
            tmp_file_path,
        )
        pr = PlateRecording.from_directory(tmp_dir)
        assert pr.get_well_indices() == tuple(range(24))

        pr.write_xlsx(
            tmp_dir, file_name=expected_excel_file, create_waveform_charts=False
        )
        assert expected_excel_file in os.listdir(tmp_dir)

        del pr  # Tanner (10/06/20): Resolve windows error with closing file when it is still open


def test_PlateRecording__creates_output_file_without_interpolation_error():
    pr = PlateRecording.from_directory(
        os.path.join(
            PATH_OF_CURRENT_FILE,
            "excel_optical_data",
            "Data_MA26_Plate2_ZIP_2020-2",
        )
    )
    pr.write_xlsx(".")
